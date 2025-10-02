"""
RAG (Retrieval-Augmented Generation) 系统核心模块

该模块包含三个主要组件：
1. DocumentProcessor: 文档处理器，负责文档内容提取、分块和向量化
2. VectorSearch: 向量搜索引擎，基于Chroma数据库进行语义相似度搜索
3. RAGGenerator: RAG生成器，集成LangChain和OpenAI进行智能问答

技术栈：
- 文档处理: PyPDF, python-docx, openpyxl
- 文本分割: LangChain RecursiveCharacterTextSplitter
- 向量化: Sentence Transformers
- 向量存储: ChromaDB
- 智能问答: LangChain + OpenAI GPT
"""

import os
import traceback
import uuid
from typing import List, Dict, Any, Optional

import chromadb
import openpyxl
import pypdf
import pytesseract
import cv2
import numpy as np
from PIL import Image
from pdf2image import convert_from_path
from chromadb.config import Settings as ChromaSettings
from docx import Document as DocxDocument
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_openai import ChatOpenAI
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser
from sentence_transformers import SentenceTransformer

from apps.models.document import Document as DocumentModel, DocumentChunk
from config import (
    CHROMA_PERSIST_DIRECTORY, CHROMA_COLLECTION, EMBEDDING_MODEL, 
    HF_HOME, HF_OFFLINE, OPENAI_API_KEY, OPENAI_BASE_URL, SIMILARITY_THRESHOLD,
    OCR_ENABLED, OCR_AUTO_FALLBACK, OCR_MIN_TEXT_LENGTH, OCR_MAX_FILE_SIZE
)

# RAG系统工具函数


def get_local_model_path(model_name: str, cache_folder: str) -> Optional[str]:
    """
    获取本地模型路径
    
    Args:
        model_name: HuggingFace模型名称
        cache_folder: 缓存文件夹路径
        
    Returns:
        本地模型路径（如果存在）或None
    """
    # HuggingFace将 '/' 转换为 '--'
    local_model_name = model_name.replace('/', '--')
    local_model_path = os.path.join(cache_folder, f"models--{local_model_name}")
    
    if os.path.exists(local_model_path):
        # 检查是否有snapshots目录
        snapshots_dir = os.path.join(local_model_path, "snapshots")
        if os.path.exists(snapshots_dir):
            # 获取最新的snapshot
            snapshots = [d for d in os.listdir(snapshots_dir) 
                        if os.path.isdir(os.path.join(snapshots_dir, d))]
            if snapshots:
                latest_snapshot = os.path.join(snapshots_dir, snapshots[0])
                return latest_snapshot
        
        # 如果没有snapshots，直接返回模型目录
        return local_model_path
    
    return None


class DocumentProcessor:
    """
    文档处理器 - 负责文档内容提取、分块和向量化
    
    主要功能：
    1. 支持多种文档格式（PDF、DOCX、Excel、TXT）
    2. 智能文本分块，保持语义完整性
    3. 生成高质量向量嵌入
    4. 与数据库和向量存储同步
    """
    
    def __init__(self):
        """
        初始化文档处理器
        
        配置组件：
        1. 文本分割器：1000字符块大小，200字符重叠
        2. 向量嵌入模型：Sentence Transformers
        3. ChromaDB向量数据库客户端
        """
        try:
            # 配置文本分割器 - 平衡块大小和语义完整性
            self.text_splitter = RecursiveCharacterTextSplitter(
                chunk_size=1000,      # 每个文本块的最大字符数
                chunk_overlap=200,    # 块之间的重叠字符数，保持上下文连续性
                length_function=len,  # 使用字符长度计算
            )
            
            # 配置HuggingFace环境变量
            os.environ["HF_HOME"] = HF_HOME
            os.environ["TRANSFORMERS_CACHE"] = HF_HOME
            os.environ["HF_HUB_CACHE"] = HF_HOME
            
            if HF_OFFLINE:
                # 离线模式配置
                os.environ["TRANSFORMERS_OFFLINE"] = "1"
                os.environ["HF_HUB_OFFLINE"] = "1"
            
            # 关闭Chroma遥测，避免网络请求和错误
            os.environ.setdefault("ANONYMIZED_TELEMETRY", "false")
            os.environ.setdefault("CHROMA_TELEMETRY", "false")
            
            # 初始化向量嵌入模型
            local_model_path = get_local_model_path(EMBEDDING_MODEL, HF_HOME)
            
            if local_model_path and HF_OFFLINE:
                self.embedding_model = SentenceTransformer(local_model_path)
            else:
                self.embedding_model = SentenceTransformer(
                    EMBEDDING_MODEL, 
                    cache_folder=HF_HOME
                )
            
            # 初始化ChromaDB客户端和集合
            os.makedirs(CHROMA_PERSIST_DIRECTORY, exist_ok=True)
            self.chroma_client = chromadb.PersistentClient(
                path=CHROMA_PERSIST_DIRECTORY, 
                settings=ChromaSettings(anonymized_telemetry=False)
            )
            self.collection = self.chroma_client.get_or_create_collection(
                name=CHROMA_COLLECTION, 
                metadata={"hnsw:space": "cosine"}  # 使用余弦相似度
            )
            
        except Exception as e:
            raise Exception(f"DocumentProcessor初始化失败: {e}")
        
    async def process_document(self, document_id: int, file_path: str, file_type: str) -> bool:
        """
        处理文档并生成向量嵌入
        
        处理流程：
        1. 提取文档内容
        2. 智能分块处理
        3. 生成向量嵌入
        4. 存储到ChromaDB
        5. 更新数据库状态
        
        Args:
            document_id: 文档ID
            file_path: 文件路径
            file_type: 文件类型
            
        Returns:
            bool: 处理是否成功
        """
        document = None
        try:
            # 1. 更新文档状态为处理中
            document = await DocumentModel.get(id=document_id)
            document.status = "processing"
            await document.save()
            
            # 2. 提取文档内容
            content = await self._extract_content(file_path, file_type)
            if not content or not content.strip():
                raise Exception("文档内容为空或无法提取")
            
            # 更新文档内容到数据库
            document.content = content
            await document.save()
            
            # 3. 智能分块处理
            chunks = self.text_splitter.split_text(content)
            if not chunks:
                raise Exception("文档分块失败")
            
            # 4. 创建分块记录
            chunk_objects = []
            for i, chunk_text in enumerate(chunks):
                chunk = await DocumentChunk.create(
                    document_id=document_id,
                    chunk_index=i,
                    content=chunk_text,
                    content_length=len(chunk_text),
                    metadata={"chunk_index": i}
                )
                chunk_objects.append(chunk)
            
            # 5. 生成向量嵌入
            embeddings = self.embedding_model.encode(chunks)
            
            # 6. 存储向量到ChromaDB
            if len(chunk_objects) > 0:
                ids = []
                metadatas = []
                vectors = []
                
                for i, (chunk, embedding) in enumerate(zip(chunk_objects, embeddings)):
                    vector_id = f"doc_{document_id}_chunk_{i}_{uuid.uuid4().hex[:8]}"
                    ids.append(vector_id)
                    metadatas.append({
                        "document_id": document_id,
                        "chunk_id": chunk.id,
                        "chunk_index": i,
                    })
                    vectors.append(embedding.tolist())
                
                # 批量添加到ChromaDB
                self.collection.add(
                    ids=ids, 
                    embeddings=vectors, 
                    metadatas=metadatas
                )
            
            # 7. 更新文档状态为完成
            document.status = "completed"
            await document.save()
            
            return True
            
        except Exception as e:
            # 更新文档状态为失败
            try:
                if document is None:
                    document = await DocumentModel.get(id=document_id)
                document.status = "failed"
                document.error_message = str(e)
                await document.save()
            except Exception:
                pass  # 忽略保存错误状态的异常
            
            return False
    
    async def _extract_content(self, file_path: str, file_type: str) -> str:
        """
        根据文件类型提取文档内容
        
        支持的文件类型：
        - PDF: 使用PyPDF提取文本
        - DOCX/DOC: 使用python-docx提取段落
        - XLSX/XLS: 使用openpyxl提取表格数据
        - TXT: 直接读取文本内容
        
        Args:
            file_path: 文件路径
            file_type: 文件类型
            
        Returns:
            str: 提取的文本内容
        """
        try:
            if not os.path.exists(file_path):
                raise Exception(f"文件不存在: {file_path}")
            
            if file_type == "pdf":
                return await self._extract_pdf_content(file_path)
            elif file_type in ["docx", "doc"]:
                return await self._extract_docx_content(file_path)
            elif file_type in ["xlsx", "xls"]:
                return await self._extract_excel_content(file_path)
            elif file_type == "txt":
                return await self._extract_txt_content(file_path)
            else:
                raise Exception(f"不支持的文件类型: {file_type}")
                
        except Exception as e:
            raise Exception(f"提取文档内容失败: {str(e)}")
    
    async def _extract_pdf_content(self, file_path: str) -> str:
        """
        智能PDF内容提取 - 根据配置和文档特征选择最佳策略
        
        处理策略：
        1. 总是先尝试文本提取（快速、准确）
        2. 根据配置和结果质量决定是否使用OCR
        3. 提供清晰的处理状态和错误信息
        
        Args:
            file_path: PDF文件路径
            
        Returns:
            str: 提取的文本内容
            
        Raises:
            Exception: 当所有提取方法都失败时
        """
        try:
            # 检查文件大小
            file_size = os.path.getsize(file_path)
            print(f"📄 开始处理PDF文档，文件大小: {file_size / 1024 / 1024:.2f}MB")
            
            # 第一步：总是先尝试文本提取
            text_content = await self._extract_pdf_text(file_path)
            text_length = len(text_content.strip()) if text_content else 0
            
            # 判断文本提取质量
            is_text_sufficient = text_length >= OCR_MIN_TEXT_LENGTH
            
            if is_text_sufficient:
                print(f"✅ PDF文本提取成功，内容长度: {text_length} 字符")
                return text_content
            
            # 第二步：决定是否使用OCR
            if not OCR_ENABLED:
                if text_content:
                    print(f"⚠️ OCR功能未启用，返回已提取的文本内容 ({text_length} 字符)")
                    return text_content
                else:
                    raise Exception("PDF无法提取文本内容，且OCR功能未启用。请启用OCR或提供文本格式的PDF。")
            
            # 检查文件大小限制
            if file_size > OCR_MAX_FILE_SIZE:
                if text_content:
                    print(f"⚠️ 文件过大 ({file_size / 1024 / 1024:.2f}MB > {OCR_MAX_FILE_SIZE / 1024 / 1024}MB)，跳过OCR处理")
                    return text_content
                else:
                    raise Exception(f"PDF文件过大 ({file_size / 1024 / 1024:.2f}MB)，无法进行OCR处理。请提供更小的文件或文本格式的PDF。")
            
            # 如果不是自动降级模式，且有一些文本内容，先返回文本内容
            if not OCR_AUTO_FALLBACK and text_content:
                print(f"ℹ️ 检测到少量文本内容 ({text_length} 字符)，OCR需手动启用")
                return text_content
            
            # 第三步：执行OCR处理
            print(f"📸 文本内容不足 ({text_length} < {OCR_MIN_TEXT_LENGTH})，开始OCR处理...")
            print("⏳ OCR处理可能需要较长时间，请耐心等待...")
            
            ocr_content = await self._extract_pdf_with_ocr(file_path)
            ocr_length = len(ocr_content.strip()) if ocr_content else 0
            
            if ocr_content and ocr_length > 10:
                print(f"✅ PDF OCR处理成功，内容长度: {ocr_length} 字符")
                return ocr_content
            
            # 最后的降级处理
            if text_content:
                print(f"⚠️ OCR处理失败，返回原始文本提取结果 ({text_length} 字符)")
                return text_content
                
            raise Exception("PDF文档处理失败：文本提取和OCR识别均未获得有效内容")
            
        except Exception as e:
            print(f"❌ PDF处理出错: {str(e)}")
            raise Exception(f"PDF内容提取失败: {str(e)}")
    
    async def _extract_pdf_text(self, file_path: str) -> str:
        """
        使用PyPDF直接提取PDF文本内容
        
        Args:
            file_path: PDF文件路径
            
        Returns:
            str: 提取的文本内容
        """
        try:
            content = ""
            with open(file_path, 'rb') as file:
                pdf_reader = pypdf.PdfReader(file)
                
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    page_text = page.extract_text()
                    if page_text.strip():  # 只添加非空页面
                        content += f"\n--- 第 {page_num} 页 ---\n"
                        content += page_text + "\n"
                        
            return content.strip()
            
        except Exception as e:
            print(f"PDF文本提取出错: {str(e)}")
            return ""
    
    async def _extract_pdf_with_ocr(self, file_path: str) -> str:
        """
        使用OCR技术提取PDF中的图片文字
        
        包含完整的依赖检查和错误处理
        
        Args:
            file_path: PDF文件路径
            
        Returns:
            str: OCR识别的文本内容
            
        Raises:
            Exception: 当OCR依赖缺失或处理失败时
        """
        try:
            # 检查OCR依赖
            self._check_ocr_dependencies()
            
            # 将PDF转换为图片
            print("🔄 正在将PDF转换为图片...")
            try:
                images = convert_from_path(file_path, dpi=200)  # 降低DPI平衡质量和性能
            except Exception as e:
                if "poppler" in str(e).lower():
                    raise Exception("缺少poppler依赖。请运行: brew install poppler (macOS) 或 apt-get install poppler-utils (Ubuntu)")
                raise Exception(f"PDF转图片失败: {str(e)}")
            
            if not images:
                raise Exception("PDF转换后未获得任何图片页面")
            
            content = ""
            total_pages = len(images)
            successful_pages = 0
            
            print(f"📄 开始OCR处理 {total_pages} 页...")
            
            for page_num, image in enumerate(images, 1):
                try:
                    print(f"🔍 处理第 {page_num}/{total_pages} 页...")
                    
                    # 图像预处理
                    processed_image = self._preprocess_image_for_ocr(image)
                    
                    # OCR识别 - 使用简单可靠的配置
                    page_text = pytesseract.image_to_string(
                        processed_image, 
                        lang='chi_sim+eng',
                        config='--oem 3 --psm 6'
                    )
                    
                    if page_text.strip():
                        content += f"\n--- 第 {page_num} 页 (OCR) ---\n"
                        content += page_text.strip() + "\n"
                        successful_pages += 1
                        
                except Exception as page_error:
                    print(f"⚠️ 第 {page_num} 页OCR处理失败: {str(page_error)}")
                    continue
            
            if successful_pages == 0:
                raise Exception("所有页面的OCR处理均失败")
            
            print(f"✅ OCR处理完成，成功处理 {successful_pages}/{total_pages} 页")
            return content.strip()
            
        except Exception as e:
            error_msg = str(e)
            if "tesseract" in error_msg.lower():
                error_msg = "缺少Tesseract OCR引擎。请运行: brew install tesseract (macOS) 或 apt-get install tesseract-ocr (Ubuntu)"
            elif "chi_sim" in error_msg.lower():
                error_msg = "缺少中文语言包。请运行: brew install tesseract-lang (macOS) 或 apt-get install tesseract-ocr-chi-sim (Ubuntu)"
            
            print(f"❌ PDF OCR处理失败: {error_msg}")
            raise Exception(f"OCR处理失败: {error_msg}")
    
    def _check_ocr_dependencies(self):
        """
        检查OCR所需的依赖是否可用
        
        Raises:
            Exception: 当依赖缺失时
        """
        try:
            # 检查pytesseract
            import pytesseract
            
            # 尝试设置tesseract路径（macOS Homebrew默认路径）
            import shutil
            tesseract_path = shutil.which('tesseract')
            if tesseract_path:
                pytesseract.pytesseract.tesseract_cmd = tesseract_path
                print(f"🔧 设置Tesseract路径: {tesseract_path}")
            
            # 检查Tesseract可执行文件
            version = pytesseract.get_tesseract_version()
            print(f"✅ Tesseract版本: {version}")
            
            # 检查支持的语言
            languages = pytesseract.get_languages()
            print(f"📋 支持的语言: {len(languages)} 种")
            
            if 'chi_sim' not in languages:
                raise Exception("Tesseract缺少中文简体语言包。请运行: brew install tesseract-lang")
            if 'eng' not in languages:
                raise Exception("Tesseract缺少英文语言包")
            
            print("✅ OCR依赖检查通过")
                
        except ImportError:
            raise Exception("pytesseract包未安装")
        except Exception as e:
            error_str = str(e).lower()
            if "tesseract is not installed" in error_str or "tesseract not found" in error_str:
                raise Exception("Tesseract OCR引擎未安装或未在PATH中")
            raise e
    
    
    def _preprocess_image_for_ocr(self, pil_image: Image.Image) -> Image.Image:
        """
        简单的图像预处理，提高OCR识别准确性
        
        Args:
            pil_image: PIL图像对象
            
        Returns:
            Image.Image: 预处理后的图像
        """
        try:
            # 简单的灰度转换
            if pil_image.mode != 'L':
                gray_image = pil_image.convert('L')
            else:
                gray_image = pil_image
            
            # 如果图像太小，稍微放大
            width, height = gray_image.size
            if width < 800 or height < 800:
                scale_factor = max(800 / width, 800 / height)
                new_width = int(width * scale_factor)
                new_height = int(height * scale_factor)
                gray_image = gray_image.resize((new_width, new_height), Image.Resampling.LANCZOS)
            
            return gray_image
            
        except Exception as e:
            print(f"⚠️ 图像预处理出错: {str(e)}")
            # 如果预处理失败，返回原图
            return pil_image
    
    async def _extract_docx_content(self, file_path: str) -> str:
        """
        提取DOCX文档内容
        
        Args:
            file_path: DOCX文件路径
            
        Returns:
            str: 提取的文本内容
        """
        try:
            doc = DocxDocument(file_path)
            content = ""
            paragraph_count = 0
            
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():  # 只添加非空段落
                    content += paragraph.text + "\n"
                    paragraph_count += 1
            
            if not content.strip():
                raise Exception("DOCX文档无法提取到有效文本内容")
                
            return content.strip()
            
        except Exception as e:
            raise Exception(f"DOCX内容提取失败: {str(e)}")
    
    async def _extract_excel_content(self, file_path: str) -> str:
        """
        提取Excel文档内容
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            str: 提取的文本内容
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            content = ""
            total_rows = 0
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content += f"\n=== 工作表: {sheet_name} ===\n"
                
                # 获取有数据的行
                rows_with_data = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None and str(cell).strip() for cell in row):
                        row_text = "\t".join([
                            str(cell).strip() if cell is not None else "" 
                            for cell in row
                        ])
                        rows_with_data.append(row_text)
                        total_rows += 1
                
                if rows_with_data:
                    content += "\n".join(rows_with_data) + "\n"
                else:
                    content += "（此工作表无数据）\n"
            
            if not content.strip():
                raise Exception("Excel文档无法提取到有效内容")
                
            return content.strip()
            
        except Exception as e:
            raise Exception(f"Excel内容提取失败: {str(e)}")
    
    async def _extract_txt_content(self, file_path: str) -> str:
        """
        提取TXT文档内容
        
        Args:
            file_path: TXT文件路径
            
        Returns:
            str: 提取的文本内容
        """
        try:
            # 尝试多种编码格式
            encodings = ['utf-8', 'gbk', 'gb2312', 'big5', 'latin1']
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as file:
                        content = file.read()
                        if content.strip():
                            return content.strip()
                except UnicodeDecodeError:
                    continue
            
            raise Exception("无法使用支持的编码格式读取文本文件")
            
        except Exception as e:
            raise Exception(f"TXT内容提取失败: {str(e)}")


class VectorSearch:
    """
    向量搜索引擎 - 基于ChromaDB的语义相似度搜索
    
    主要功能：
    1. 语义相似度搜索
    2. 向量数据管理
    3. 搜索结果排序和过滤
    4. 与数据库数据关联
    """
    
    def __init__(self):
        """
        初始化向量搜索引擎
        
        配置组件：
        1. 向量嵌入模型（与DocumentProcessor保持一致）
        2. ChromaDB客户端和集合
        """
        try:
            # 配置HuggingFace环境
            os.environ["HF_HOME"] = HF_HOME
            os.environ["TRANSFORMERS_CACHE"] = HF_HOME
            os.environ["HF_HUB_CACHE"] = HF_HOME
            
            if HF_OFFLINE:
                os.environ["TRANSFORMERS_OFFLINE"] = "1"
                os.environ["HF_HUB_OFFLINE"] = "1"
            
            # 关闭Chroma遥测
            os.environ.setdefault("ANONYMIZED_TELEMETRY", "false")
            os.environ.setdefault("CHROMA_TELEMETRY", "false")
            
            # 初始化向量嵌入模型（与DocumentProcessor使用相同模型）
            local_model_path = get_local_model_path(EMBEDDING_MODEL, HF_HOME)
            
            if local_model_path and HF_OFFLINE:
                self.embedding_model = SentenceTransformer(local_model_path)
            else:
                self.embedding_model = SentenceTransformer(
                    EMBEDDING_MODEL, 
                    cache_folder=HF_HOME
                )
            
            # 初始化ChromaDB客户端
            self.chroma_client = chromadb.PersistentClient(
                path=CHROMA_PERSIST_DIRECTORY, 
                settings=ChromaSettings(anonymized_telemetry=False)
            )
            self.collection = self.chroma_client.get_or_create_collection(
                name=CHROMA_COLLECTION, 
                metadata={"hnsw:space": "cosine"}
            )
            
        except Exception as e:
            raise Exception(f"VectorSearch初始化失败: {e}")
    
    async def search_similar_chunks(self, query: str, top_k: int = 5, use_threshold: bool = True) -> List[Dict[str, Any]]:
        """
        搜索语义相似的文档块
        
        搜索流程：
        1. 将查询文本转换为向量
        2. 在ChromaDB中进行相似度搜索
        3. 从数据库获取完整的文档和块信息
        4. 计算相似度分数并排序
        
        Args:
            query: 查询文本
            top_k: 返回结果数量
            use_threshold: 是否使用相似度阈值过滤
            
        Returns:
            List[Dict]: 相似文档块列表，包含文档、块和相似度信息
        """
        try:
            if not query or not query.strip():
                return []
            
            # 1. 生成查询向量
            query_embedding = self.embedding_model.encode([query.strip()])[0]
            
            # 2. 在ChromaDB中搜索相似向量
            results = self.collection.query(
                query_embeddings=[query_embedding.tolist()], 
                n_results=min(top_k, 20)  # 限制最大搜索数量
            )
            
            # 3. 解析搜索结果
            ids = results.get('ids', [[]])[0]
            metadatas = results.get('metadatas', [[]])[0]
            distances = results.get('distances', [[]])[0] or []
            
            # 4. 从数据库获取完整信息并构建结果
            all_similarities = []  # 存储所有结果
            filtered_similarities = []  # 存储过滤后的结果
            
            for i, (vector_id, metadata) in enumerate(zip(ids, metadatas)):
                try:
                    # 获取文档块信息
                    chunk_id = metadata.get('chunk_id')
                    if not chunk_id:
                        continue
                    
                    chunk = await DocumentChunk.get_or_none(id=chunk_id)
                    if not chunk:
                        continue
                    
                    # 获取关联的文档
                    document = await chunk.document
                    if not document:
                        continue
                    
                    # 计算相似度分数（距离转换为相似度）
                    distance = distances[i] if i < len(distances) else 1.0
                    similarity = max(0.0, 1.0 - float(distance))
                    
                    result_item = {
                        'vector_id': vector_id,
                        'similarity': similarity,
                        'chunk': chunk,
                        'document': document,
                        'metadata': metadata,
                        'above_threshold': similarity >= SIMILARITY_THRESHOLD
                    }
                    
                    all_similarities.append(result_item)
                    
                    # 如果使用阈值过滤，只保留相似度大于阈值的结果
                    if not use_threshold or similarity >= SIMILARITY_THRESHOLD:
                        filtered_similarities.append(result_item)
                    
                except Exception as e:
                    print(f"处理搜索结果项失败 (chunk_id: {metadata.get('chunk_id', 'unknown')}): {e}")
                    continue
            
            # 5. 智能选择返回结果
            if use_threshold and filtered_similarities:
                # 有超过阈值的结果，返回过滤后的结果
                filtered_similarities.sort(key=lambda x: x['similarity'], reverse=True)
                return filtered_similarities[:top_k]
            elif use_threshold and not filtered_similarities and all_similarities:
                # 没有超过阈值的结果，但有搜索结果，返回最相似的几个并标记
                all_similarities.sort(key=lambda x: x['similarity'], reverse=True)
                # 取前几个最相似的结果，但标记为低相似度
                return all_similarities[:min(top_k, 3)]  # 最多返回3个低相似度结果
            else:
                # 不使用阈值过滤，返回所有结果
                all_similarities.sort(key=lambda x: x['similarity'], reverse=True)
                return all_similarities[:top_k]
            
        except Exception as e:
            print(f"搜索相似文档块失败: {e}")
            import traceback
            traceback.print_exc()
            return []

    async def delete_document_vectors(self, document_id: int):
        """
        删除指定文档的所有向量数据
        
        Args:
            document_id: 文档ID
        """
        try:
            # 通过metadata条件删除
            self.collection.delete(where={"document_id": document_id})
            
        except Exception as e:
            raise Exception(f"删除文档 {document_id} 向量数据失败: {e}")

    async def count_vectors(self) -> int:
        """
        统计向量数据库中的向量总数
        
        Returns:
            int: 向量总数
        """
        try:
            count = self.collection.count()
            return int(count) if isinstance(count, (int, float)) else 0
            
        except Exception:
            return 0


class RAGGenerator:
    """
    RAG生成器 - 检索增强生成系统
    
    主要功能：
    1. 集成向量搜索和LLM生成
    2. 基于检索到的文档内容生成智能回答
    3. 支持多种回答模式（LLM智能回答 / 简单回答）
    4. 提供上下文感知的问答体验
    """
    
    def __init__(self):
        """
        初始化RAG生成器
        
        组件：
        1. VectorSearch: 向量搜索引擎
        2. LangChain + OpenAI: 智能问答链（可选）
        3. 备用简单回答模式
        """
        try:
            # 初始化向量搜索引擎
            self.vector_search = VectorSearch()
            
            # 初始化LLM组件
            self.llm = None
            self.chain = None
            self.llm_available = False
            
            # 检查OpenAI配置并初始化LLM
            if OPENAI_API_KEY and OPENAI_API_KEY.strip():
                try:
                    # 创建OpenAI客户端
                    self.llm = ChatOpenAI(
                        api_key=OPENAI_API_KEY,
                        base_url=OPENAI_BASE_URL,
                        temperature=0.1,  # 较低的温度确保回答的一致性
                        model="gpt-3.5-turbo",
                        max_tokens=2000  # 限制回答长度
                    )
                    
                    # 创建系统提示模板
                    system_template = """你是一个专业的文档问答助手。请基于提供的文档内容准确回答用户的问题。

回答要求：
1. 严格基于提供的文档内容回答，不要添加文档中没有的信息
2. 如果文档中没有相关信息，请明确说明"根据提供的文档内容，无法找到相关信息"
3. 回答要准确、详细且有条理，使用清晰的段落结构：
   - 使用标题和子标题组织内容
   - 重要信息用**粗体**标记
   - 使用项目符号(•)或数字列表展示要点
   - 每个段落专注一个主题，段落间留空行
   - 复杂内容使用表格或结构化格式
4. 可以引用具体的文档名称和关键内容片段，格式：「文档名：具体内容」
5. 如果有多个文档提供了相关信息，请综合分析并标明信息来源
6. 用中文回答，语言要专业但易懂，适当使用emoji增强可读性

提供的文档内容：
{context}"""

                    human_template = "问题：{question}"
                    
                    # 创建聊天提示模板
                    chat_prompt = ChatPromptTemplate([
                        ("system", system_template),
                        ("human", human_template),
                    ])
                    
                    # 创建输出解析器
                    output_parser = StrOutputParser()
                    
                    # 创建LangChain处理链
                    self.chain = chat_prompt | self.llm | output_parser
                    self.llm_available = True
                    
                except Exception:
                    self.llm = None
                    self.chain = None
                    self.llm_available = False
            else:
                self.llm_available = False
                
        except Exception as e:
            raise Exception(f"RAGGenerator初始化失败: {e}")
    
    async def generate_answer(self, query: str, context_chunks: List[Dict[str, Any]]) -> str:
        """
        基于检索到的文档内容生成智能回答
        
        回答策略：
        1. 优先使用LLM智能回答（如果可用）
        2. 降级到简单回答模式（备用方案）
        3. 提供上下文信息和来源引用
        
        Args:
            query: 用户问题
            context_chunks: 检索到的相关文档块
            
        Returns:
            str: 生成的回答
        """
        try:
            # 检查是否有相关文档
            if not context_chunks:
                return "抱歉，没有找到相关的文档内容来回答您的问题。请确保已上传相关文档，或尝试使用不同的关键词重新提问。"
            
            # 构建上下文信息
            context_parts = []
            for i, chunk in enumerate(context_chunks, 1):
                doc_name = chunk['document'].original_filename or chunk['document'].filename
                content = chunk['chunk'].content
                similarity = chunk['similarity']
                
                # 格式化文档片段
                context_part = f"""文档 {i}: {doc_name}
相似度: {similarity:.2f}
内容: {content}"""
                context_parts.append(context_part)
            
            context = "\n\n" + "="*50 + "\n\n".join(context_parts)
            
            # 选择回答模式
            if self.llm_available and self.chain:
                try:
                    answer = await self._llm_answer(query, context)
                    
                    # 不再自动添加来源信息，由前端决定是否显示
                    return answer
                    
                except Exception:
                    return self._simple_answer(query, context_chunks)
            else:
                return self._simple_answer(query, context_chunks)
                
        except Exception as e:
            return f"抱歉，生成回答时出现错误: {str(e)}"
    
    async def generate_answer_stream(self, query: str, context_chunks: List[Dict[str, Any]]):
        """
        流式生成回答
        
        Args:
            query: 用户问题
            context_chunks: 文档上下文块列表
            
        Yields:
            str: 流式生成的文本块
        """
        try:
            # 构建上下文
            context_parts = []
            for i, chunk in enumerate(context_chunks, 1):
                doc_name = chunk.get('document_name', '未知文档')
                content = chunk.get('content', '')
                context_parts.append(f"文档{i}: {doc_name}\n内容: {content}")
            
            context = "\n\n" + "="*50 + "\n\n".join(context_parts)
            
            # 选择回答模式
            if self.llm_available and self.chain:
                try:
                    async for chunk in self._llm_answer_stream(query, context):
                        yield chunk
                except Exception as e:
                    print(f"LLM流式生成失败: {e}")
                    # 降级到简单回答
                    simple_answer = self._simple_answer(query, context_chunks)
                    yield simple_answer
            else:
                # 非LLM模式，直接返回简单回答
                simple_answer = self._simple_answer(query, context_chunks)
                yield simple_answer
                
        except Exception as e:
            yield f"抱歉，生成回答时出现错误: {str(e)}"

    async def _llm_answer_stream(self, query: str, context: str):
        """
        使用LangChain链式流式生成智能回答
        
        Args:
            query: 用户问题
            context: 文档上下文
            
        Yields:
            str: 流式生成的文本块
        """
        try:
            # 使用已经创建好的chain进行流式调用
            if self.chain:
                async for chunk in self.chain.astream({
                    "question": query,
                    "context": context
                }):
                    if chunk:
                        yield chunk
            else:
                raise Exception("LangChain处理链未初始化")
                    
        except Exception as e:
            traceback.print_exc()
            raise Exception(f"LLM链式流式调用失败: {str(e)}")
    
    async def _llm_answer(self, query: str, context: str) -> str:
        """
        使用LLM生成智能回答
        
        Args:
            query: 用户问题
            context: 文档上下文
            
        Returns:
            str: LLM生成的回答
        """
        try:
            # 调用LangChain处理链
            response = await self.chain.ainvoke({
                "question": query,
                "context": context
            })
            
            if not response or not response.strip():
                raise Exception("LLM返回空回答")
            
            return response.strip()
            
        except Exception as e:
            raise Exception(f"LLM调用失败: {str(e)}")
    
    def _simple_answer(self, query: str, context_chunks: List[Dict[str, Any]]) -> str:
        """
        简单回答模式 - 基于关键词匹配的备用方案
        
        Args:
            query: 用户问题
            context_chunks: 相关文档块
            
        Returns:
            str: 简单模式生成的回答
        """
        try:
            if not context_chunks:
                return "抱歉，没有找到相关的文档内容来回答您的问题。"
            
            # 获取最相关的文档片段
            best_chunk = context_chunks[0]
            document_name = best_chunk['document'].original_filename or best_chunk['document'].filename
            content = best_chunk['chunk'].content
            similarity = best_chunk['similarity']
            
            # 构建简单回答（不包含参考来源）
            answer = f"""基于文档《{document_name}》中的相关内容：

{content}

💡 **提示**：当前使用简单回答模式。配置OpenAI API Key后可获得更智能的回答。"""
            
            return answer
            
        except Exception:
            return "抱歉，生成回答时出现错误。"
    
    def _build_sources_info(self, context_chunks: List[Dict[str, Any]]) -> str:
        """
        构建来源信息
        
        Args:
            context_chunks: 文档块列表
            
        Returns:
            str: 格式化的来源信息
        """
        try:
            sources = []
            for i, chunk in enumerate(context_chunks, 1):
                doc_name = chunk['document'].original_filename or chunk['document'].filename
                similarity = chunk['similarity']
                sources.append(f"• {doc_name} (相似度: {similarity:.1%})")
            
            sources_text = "\n".join(sources)
            return f"""---
📋 **参考来源**：
{sources_text}

💡 基于 {len(context_chunks)} 个相关文档片段生成此回答"""
            
        except Exception:
            return "---\n📋 **参考来源**：信息获取失败"


# 全局实例 - 单例模式，确保整个应用使用相同的实例
try:
    # 文档处理器实例
    document_processor = DocumentProcessor()
    
    # 向量搜索实例
    vector_search = VectorSearch()
    
    # RAG生成器实例
    rag_generator = RAGGenerator()
    
except Exception as e:
    raise Exception(f"RAG系统初始化失败: {e}")
