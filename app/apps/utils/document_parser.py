"""
æ–‡æ¡£è§£æå™¨æ¨¡å— - ä¸“é—¨è´Ÿè´£æ–‡æ¡£å†…å®¹æå–å’Œè§£æ

è¯¥æ¨¡å—æä¾›ç»Ÿä¸€çš„æ–‡æ¡£è§£ææ¥å£ï¼Œæ”¯æŒå¤šç§æ–‡ä»¶æ ¼å¼ï¼š
- PDF: PyMuPDFLoaderï¼ˆä¼˜å…ˆï¼‰â†’ PyPDFLoader â†’ OCRï¼ˆæ‰«æç‰ˆPDFï¼‰
- DOCX/DOC: Docx2txtLoader
- XLSX/XLS: SimpleExcelLoaderï¼ˆè‡ªå®šä¹‰ï¼Œä½¿ç”¨ openpyxlï¼‰
- TXT: TextLoader
- MD: UnstructuredMarkdownLoader

æŠ€æœ¯æ ˆï¼š
- LangChainæ–‡æ¡£åŠ è½½å™¨
- openpyxl Excelå¤„ç†ï¼ˆç¦»çº¿ï¼Œæ— ç½‘ç»œä¾èµ–ï¼‰
- OCRå¼•æ“ï¼ˆEasyOCRï¼‰
- å›¾åƒé¢„å¤„ç†å’Œä¼˜åŒ–
"""

import asyncio
import os
import shutil
from datetime import datetime

import openpyxl
from PIL import Image
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import (
    PyPDFLoader,
    PyMuPDFLoader,
    Docx2txtLoader,
    TextLoader,
    UnstructuredMarkdownLoader
)
from langchain_core.documents import Document
from pdf2image import convert_from_path

from apps.models.document import Document as DocumentModel, DocumentChunk
from apps.utils.ocr_engines import get_ocr_engine
from apps.utils.rag_helper import vector_search
from config import OCR_ENABLED, OCR_USE_GPU, HF_HOME, HF_OFFLINE


class SimpleExcelLoader:
    """
    ç®€å•çš„ Excel åŠ è½½å™¨ - é¿å… UnstructuredExcelLoader çš„ NLTK ä¾èµ–
    
    ä½¿ç”¨ openpyxl ç›´æ¥è¯»å– Excelï¼Œå…¼å®¹ LangChain æ¥å£
    """

    def __init__(self, file_path: str):
        self.file_path = file_path

    def load(self):
        """åŠ è½½ Excel å¹¶è¿”å› LangChain Document åˆ—è¡¨"""

        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)
            documents = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []

                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    if any(val.strip() for val in row_values):
                        rows.append("\t".join(row_values))

                if rows:
                    content = f"å·¥ä½œè¡¨: {sheet_name}\n\n" + "\n".join(rows)
                    doc = Document(
                        page_content=content,
                        metadata={"source": self.file_path, "sheet_name": sheet_name}
                    )
                    documents.append(doc)

            wb.close()
            return documents

        except Exception as e:
            raise Exception(f"Excel è¯»å–å¤±è´¥: {str(e)}")


class DocumentParser:
    """
    æ–‡æ¡£è§£æå™¨ - ä¸“é—¨è´Ÿè´£æ–‡æ¡£å†…å®¹æå–
    
    ä¸»è¦åŠŸèƒ½ï¼š
    1. ä½¿ç”¨LangChainæ–‡æ¡£åŠ è½½å™¨å¤„ç†å¤šç§æ ¼å¼
    2. OCRå¤„ç†æ‰«æç‰ˆPDF
    3. å›¾åƒé¢„å¤„ç†å’Œä¼˜åŒ–
    4. ç»Ÿä¸€çš„æ–‡æ¡£è§£ææ¥å£
    """

    def __init__(self):
        """
        åˆå§‹åŒ–æ–‡æ¡£è§£æå™¨
        
        é…ç½®ç»„ä»¶ï¼š
        1. OCRå¼•æ“ï¼šEasyOCRå®ä¾‹ï¼ˆå¦‚æœå¯ç”¨ï¼‰
        2. ä¾èµ–æ£€æŸ¥å·¥å…·
        """
        try:
            # åˆå§‹åŒ–OCRå¼•æ“ï¼ˆå¦‚æœå¯ç”¨ï¼‰
            self.ocr_engine = None
            if OCR_ENABLED:
                try:
                    self.ocr_engine = get_ocr_engine(use_gpu=OCR_USE_GPU)
                    print("âœ… OCRå¼•æ“åˆå§‹åŒ–å®Œæˆ")
                except Exception as e:
                    print(f"âš ï¸ OCRå¼•æ“åˆå§‹åŒ–å¤±è´¥: {str(e)}")
                    self.ocr_engine = None

        except Exception as e:
            raise Exception(f"DocumentParseråˆå§‹åŒ–å¤±è´¥: {e}")

    async def extract_content(self, file_path: str, file_type: str) -> str:
        """
        æå–æ–‡æ¡£å†…å®¹çš„ä¸»å…¥å£æ–¹æ³•
        
        å¤„ç†æµç¨‹ï¼š
        1. æ ¹æ®æ–‡ä»¶ç±»å‹é€‰æ‹©åˆé€‚çš„åŠ è½½å™¨
        2. ä½¿ç”¨LangChainåŠ è½½å™¨æå–å†…å®¹
        3. å¦‚æœå¤±è´¥ä¸”æ˜¯PDFï¼Œå°è¯•OCRå¤„ç†
        4. è¿”å›æå–çš„æ–‡æœ¬å†…å®¹
        
        Args:
            file_path: æ–‡ä»¶è·¯å¾„
            file_type: æ–‡ä»¶ç±»å‹
            
        Returns:
            str: æå–çš„æ–‡æœ¬å†…å®¹
        """
        try:
            if not os.path.exists(file_path):
                raise Exception(f"æ–‡ä»¶ä¸å­˜åœ¨: {file_path}")

            print(f"ğŸ“„ å¼€å§‹è§£æ {file_type.upper()} æ–‡æ¡£: {os.path.basename(file_path)}")

            # æ ¹æ®æ–‡ä»¶ç±»å‹é€‰æ‹©åˆé€‚çš„åŠ è½½å™¨
            loaders = self._get_loaders(file_path, file_type)

            # åŠ è½½æ–‡æ¡£å¹¶åˆå¹¶å†…å®¹
            texts = []
            for loader in loaders:
                try:
                    documents = loader.load()
                    texts.extend(documents)
                except Exception as e:
                    print(f"âš ï¸ åŠ è½½å™¨å¤±è´¥: {str(e)}")
                    continue

            if not texts:
                raise Exception("æ— æ³•åŠ è½½ä»»ä½•æ–‡æ¡£å†…å®¹")

            # åˆå¹¶æ‰€æœ‰æ–‡æ¡£å†…å®¹
            content = "\n\n".join([doc.page_content for doc in texts if doc.page_content.strip()])

            if not content.strip():
                raise Exception("æ–‡æ¡£å†…å®¹ä¸ºç©º")

            print(f"âœ… æ–‡æ¡£è§£ææˆåŠŸï¼Œæå–å†…å®¹é•¿åº¦: {len(content)} å­—ç¬¦")
            return content.strip()

        except Exception as e:
            print(f"âŒ æ–‡æ¡£è§£æå¤±è´¥: {str(e)}")
            # å¦‚æœæ˜¯PDFä¸”å¤±è´¥ï¼Œå°è¯•OCRå¤„ç†
            if file_type == "pdf":
                print("ğŸ”„ å°è¯•OCRå¤„ç†æ‰«æç‰ˆPDF...")
                try:
                    return await self._extract_pdf_with_ocr(file_path)
                except Exception as ocr_e:
                    raise Exception(f"æ‰€æœ‰PDFå¤„ç†æ–¹æ³•éƒ½å¤±è´¥: LangChain({str(e)}), OCR({str(ocr_e)})")
            else:
                raise Exception(f"æ–‡æ¡£å†…å®¹æå–å¤±è´¥: {str(e)}")

    def _get_loaders(self, file_path: str, file_type: str) -> list:
        """
        æ ¹æ®æ–‡ä»¶ç±»å‹è·å–åˆé€‚çš„LangChainåŠ è½½å™¨
        
        Args:
            file_path: æ–‡ä»¶è·¯å¾„
            file_type: æ–‡ä»¶ç±»å‹
            
        Returns:
            list: åŠ è½½å™¨åˆ—è¡¨
        """
        loaders = []

        if file_type == "pdf":
            # PDFä¼˜å…ˆä½¿ç”¨PyMuPDFLoaderï¼ˆæ›´å¿«æ›´å‡†ç¡®ï¼‰
            try:
                loaders.append(PyMuPDFLoader(file_path))
            except Exception as _:
                loaders.append(PyPDFLoader(file_path))
        elif file_type in ["docx", "doc"]:
            loaders.append(Docx2txtLoader(file_path))
        elif file_type in ["xlsx", "xls"]:
            # ä½¿ç”¨è‡ªå®šä¹‰åŠ è½½å™¨ï¼Œå®Œå…¨é¿å¼€ UnstructuredExcelLoader çš„ NLTK ä¾èµ–
            loaders.append(SimpleExcelLoader(file_path))
        elif file_type == "txt":
            loaders.append(TextLoader(file_path, encoding='utf-8'))
        elif file_type == "md":
            loaders.append(UnstructuredMarkdownLoader(file_path))
        else:
            raise Exception(f"ä¸æ”¯æŒçš„æ–‡ä»¶ç±»å‹: {file_type}")

        return loaders

    async def _extract_pdf_with_ocr(self, file_path: str) -> str:
        """
        ä½¿ç”¨OCRæŠ€æœ¯æå–PDFä¸­çš„å›¾ç‰‡æ–‡å­—
        
        åŒ…å«å®Œæ•´çš„ä¾èµ–æ£€æŸ¥å’Œé”™è¯¯å¤„ç†
        
        Args:
            file_path: PDFæ–‡ä»¶è·¯å¾„
            
        Returns:
            str: OCRè¯†åˆ«çš„æ–‡æœ¬å†…å®¹
            
        Raises:
            Exception: å½“OCRä¾èµ–ç¼ºå¤±æˆ–å¤„ç†å¤±è´¥æ—¶
        """
        try:
            # æ£€æŸ¥OCRä¾èµ–
            self._check_ocr_dependencies()

            # å°†PDFè½¬æ¢ä¸ºå›¾ç‰‡ï¼ˆå¼‚æ­¥å¤„ç†ï¼Œé¿å…é˜»å¡ï¼‰
            print("ğŸ”„ æ­£åœ¨å°†PDFè½¬æ¢ä¸ºå›¾ç‰‡...")
            try:
                # å¼‚æ­¥å¤„ç†PDFè½¬å›¾ç‰‡ï¼Œé¿å…é˜»å¡
                images = await asyncio.get_event_loop().run_in_executor(
                    None,
                    lambda: convert_from_path(file_path, dpi=200)  # é™ä½DPIå¹³è¡¡è´¨é‡å’Œæ€§èƒ½
                )
            except Exception as e:
                if "poppler" in str(e).lower():
                    raise Exception(
                        "ç¼ºå°‘popplerä¾èµ–ã€‚è¯·è¿è¡Œ: brew install poppler (macOS) æˆ– apt-get install poppler-utils (Ubuntu)")
                raise Exception(f"PDFè½¬å›¾ç‰‡å¤±è´¥: {str(e)}")

            if not images:
                raise Exception("PDFè½¬æ¢åæœªè·å¾—ä»»ä½•å›¾ç‰‡é¡µé¢")

            content = ""
            total_pages = len(images)
            successful_pages = 0

            print(f"ğŸ“„ å¼€å§‹OCRå¤„ç† {total_pages} é¡µ...")

            # ä½¿ç”¨å¼‚æ­¥å¤„ç†ï¼Œé¿å…é˜»å¡ä¸»è¿›ç¨‹
            # é™åˆ¶å¹¶å‘æ•°é‡ï¼Œé¿å…èµ„æºè€—å°½
            semaphore = asyncio.Semaphore(2)  # æœ€å¤šåŒæ—¶å¤„ç†2é¡µ

            # æ·»åŠ è¶…æ—¶æ§åˆ¶ï¼Œé¿å…å•ä¸ªé¡µé¢å¤„ç†æ—¶é—´è¿‡é•¿
            TIMEOUT_PER_PAGE = 60  # æ¯é¡µæœ€å¤š60ç§’

            async def process_single_page(page_num, image):
                async with semaphore:
                    try:
                        # æ˜¾ç¤ºå¤„ç†è¿›åº¦
                        progress = (page_num - 1) / total_pages * 100
                        print(f"ğŸ” å¤„ç†ç¬¬ {page_num}/{total_pages} é¡µ... ({progress:.1f}%)")

                        # å›¾åƒé¢„å¤„ç†ï¼ˆå¿«é€Ÿæ“ä½œï¼Œä¸éœ€è¦å¼‚æ­¥ï¼‰
                        processed_image = self._preprocess_image_for_ocr(image)

                        # OCRè¯†åˆ« - ä½¿ç”¨å·²åˆå§‹åŒ–çš„OCRå¼•æ“
                        if self.ocr_engine is None:
                            raise Exception("OCRå¼•æ“æœªåˆå§‹åŒ–")

                        # å¼‚æ­¥å¤„ç†OCRï¼Œé¿å…é˜»å¡ä¸»è¿›ç¨‹
                        # ä½¿ç”¨çº¿ç¨‹æ± æ‰§è¡Œå™¨ï¼Œè®©OCRåœ¨ç‹¬ç«‹çº¿ç¨‹ä¸­è¿è¡Œ
                        # æ·»åŠ è¶…æ—¶æ§åˆ¶
                        page_text = await asyncio.wait_for(
                            asyncio.get_event_loop().run_in_executor(
                                None,
                                self.ocr_engine.extract_text,
                                processed_image
                            ),
                            timeout=TIMEOUT_PER_PAGE
                        )

                        # æ¸…ç†å›¾åƒèµ„æº
                        del processed_image

                        return page_num, page_text, None

                    except asyncio.TimeoutError:
                        print(f"âš ï¸ ç¬¬ {page_num} é¡µOCRå¤„ç†è¶…æ—¶")
                        return page_num, None, "å¤„ç†è¶…æ—¶"
                    except Exception as page_error:
                        print(f"âš ï¸ ç¬¬ {page_num} é¡µOCRå¤„ç†å¤±è´¥: {str(page_error)}")
                        return page_num, None, str(page_error)

            # åˆ›å»ºæ‰€æœ‰é¡µé¢çš„å¤„ç†ä»»åŠ¡
            tasks = [process_single_page(page_num, image) for page_num, image in enumerate(images, 1)]

            # å¹¶å‘å¤„ç†æ‰€æœ‰é¡µé¢ï¼Œä½†é™åˆ¶å¹¶å‘æ•°é‡
            # ä½¿ç”¨ as_completed ä½†ä¿æŒé¡ºåº
            results = {}
            for task in asyncio.as_completed(tasks):
                page_num, page_text, error = await task
                results[page_num] = (page_text, error)

            # æŒ‰é¡µé¢é¡ºåºå¤„ç†ç»“æœ
            for page_num in range(1, total_pages + 1):
                if page_num in results:
                    page_text, error = results[page_num]
                    if error is None and page_text and page_text.strip():
                        content += f"\n--- ç¬¬ {page_num} é¡µ (OCR) ---\n"
                        content += page_text.strip() + "\n"
                        successful_pages += 1

            if successful_pages == 0:
                raise Exception("æ‰€æœ‰é¡µé¢çš„OCRå¤„ç†å‡å¤±è´¥")

            print(f"âœ… OCRå¤„ç†å®Œæˆï¼ŒæˆåŠŸå¤„ç† {successful_pages}/{total_pages} é¡µ")
            return content.strip()

        except Exception as e:
            error_msg = str(e)
            if "tesseract" in error_msg.lower():
                error_msg = "ç¼ºå°‘Tesseract OCRå¼•æ“ã€‚è¯·è¿è¡Œ: brew install tesseract (macOS) æˆ– apt-get install tesseract-ocr (Ubuntu)"
            elif "chi_sim" in error_msg.lower():
                error_msg = "ç¼ºå°‘ä¸­æ–‡è¯­è¨€åŒ…ã€‚è¯·è¿è¡Œ: brew install tesseract-lang (macOS) æˆ– apt-get install tesseract-ocr-chi-sim (Ubuntu)"

            print(f"âŒ PDF OCRå¤„ç†å¤±è´¥: {error_msg}")
            raise Exception(f"OCRå¤„ç†å¤±è´¥: {error_msg}")

    def _check_ocr_dependencies(self):
        """
        æ£€æŸ¥OCRæ‰€éœ€çš„ä¾èµ–æ˜¯å¦å¯ç”¨
        
        Raises:
            Exception: å½“ä¾èµ–ç¼ºå¤±æ—¶
        """
        try:
            # æ£€æŸ¥popplerå·¥å…·ï¼ˆPDFè½¬å›¾ç‰‡éœ€è¦ï¼‰
            poppler_path = shutil.which('pdftoppm')
            if not poppler_path:
                raise Exception(
                    "ç¼ºå°‘popplerå·¥å…·ï¼Œè¯·å®‰è£…: brew install poppler (macOS) æˆ– sudo apt-get install poppler-utils (Ubuntu)")
            print("âœ… OCRä¾èµ–æ£€æŸ¥é€šè¿‡")
        except Exception as e:
            raise e

    def _preprocess_image_for_ocr(self, pil_image: Image.Image) -> Image.Image:
        """
        ç®€å•çš„å›¾åƒé¢„å¤„ç†ï¼Œæé«˜OCRè¯†åˆ«å‡†ç¡®æ€§
        
        Args:
            pil_image: PILå›¾åƒå¯¹è±¡
            
        Returns:
            Image.Image: é¢„å¤„ç†åçš„å›¾åƒ
        """
        try:
            # ç®€å•çš„ç°åº¦è½¬æ¢
            if pil_image.mode != 'L':
                gray_image = pil_image.convert('L')
            else:
                gray_image = pil_image

            # å¦‚æœå›¾åƒå¤ªå°ï¼Œç¨å¾®æ”¾å¤§
            width, height = gray_image.size
            if width < 800 or height < 800:
                scale_factor = max(800 / width, 800 / height)
                new_width = int(width * scale_factor)
                new_height = int(height * scale_factor)
                gray_image = gray_image.resize((new_width, new_height), Image.Resampling.LANCZOS)

            return gray_image

        except Exception as e:
            print(f"âš ï¸ å›¾åƒé¢„å¤„ç†å‡ºé”™: {str(e)}")
            # å¦‚æœé¢„å¤„ç†å¤±è´¥ï¼Œè¿”å›åŸå›¾
            return pil_image


class DocumentProcessor:
    """
    æ–‡æ¡£å¤„ç†å™¨ - è´Ÿè´£æ–‡æ¡£åˆ†å—å’Œå‘é‡åŒ–
    
    ä¸»è¦åŠŸèƒ½ï¼š
    1. æ™ºèƒ½æ–‡æœ¬åˆ†å—ï¼Œä¿æŒè¯­ä¹‰å®Œæ•´æ€§
    2. ç”Ÿæˆé«˜è´¨é‡å‘é‡åµŒå…¥
    3. ä¸æ•°æ®åº“å’Œå‘é‡å­˜å‚¨åŒæ­¥
    4. åè°ƒæ–‡æ¡£è§£æå™¨å’Œå‘é‡åŒ–æµç¨‹
    
    å¤„ç†æµç¨‹ï¼š
    1. ä½¿ç”¨DocumentParseræå–æ–‡æ¡£å†…å®¹
    2. æ™ºèƒ½åˆ†å—å¤„ç†
    3. ç”Ÿæˆå‘é‡åµŒå…¥
    4. å­˜å‚¨åˆ°LangChainå‘é‡æ•°æ®åº“
    """

    def __init__(self):
        """
        åˆå§‹åŒ–æ–‡æ¡£å¤„ç†å™¨
        
        é…ç½®ç»„ä»¶ï¼š
        1. æ–‡æœ¬åˆ†å‰²å™¨ï¼š1000å­—ç¬¦å—å¤§å°ï¼Œ200å­—ç¬¦é‡å 
        2. æ–‡æ¡£è§£æå™¨ï¼šç‹¬ç«‹çš„DocumentParserå®ä¾‹
        3. LangChainå‘é‡å­˜å‚¨ï¼šè‡ªåŠ¨å¤„ç†å‘é‡åŒ–
        """
        try:
            # é…ç½®æ–‡æœ¬åˆ†å‰²å™¨ - å¹³è¡¡å—å¤§å°å’Œè¯­ä¹‰å®Œæ•´æ€§
            self.text_splitter = RecursiveCharacterTextSplitter(
                chunk_size=1000,  # æ¯ä¸ªæ–‡æœ¬å—çš„æœ€å¤§å­—ç¬¦æ•°
                chunk_overlap=200,  # å—ä¹‹é—´çš„é‡å å­—ç¬¦æ•°ï¼Œä¿æŒä¸Šä¸‹æ–‡è¿ç»­æ€§
                length_function=len,  # ä½¿ç”¨å­—ç¬¦é•¿åº¦è®¡ç®—
            )

            # é…ç½®HuggingFaceç¯å¢ƒå˜é‡
            os.environ["HF_HOME"] = HF_HOME
            os.environ["TRANSFORMERS_CACHE"] = HF_HOME
            os.environ["HF_HUB_CACHE"] = HF_HOME

            if HF_OFFLINE:
                # ç¦»çº¿æ¨¡å¼é…ç½®
                os.environ["TRANSFORMERS_OFFLINE"] = "1"
                os.environ["HF_HUB_OFFLINE"] = "1"
        except Exception as e:
            raise Exception(f"DocumentProcessoråˆå§‹åŒ–å¤±è´¥: {e}")

    async def process_document(self, document_id: int, file_path: str, file_type: str) -> bool:
        """
        å¤„ç†æ–‡æ¡£å¹¶ç”Ÿæˆå‘é‡åµŒå…¥
        
        å¤„ç†æµç¨‹ï¼š
        1. æå–æ–‡æ¡£å†…å®¹
        2. æ™ºèƒ½åˆ†å—å¤„ç†
        3. å­˜å‚¨åˆ°ChromaDB
        4. æ›´æ–°æ•°æ®åº“çŠ¶æ€
        
        Args:
            document_id: æ–‡æ¡£ID
            file_path: æ–‡ä»¶è·¯å¾„
            file_type: æ–‡ä»¶ç±»å‹
            
        Returns:
            bool: å¤„ç†æ˜¯å¦æˆåŠŸ
        """
        document = None
        try:
            # 1. æ›´æ–°æ–‡æ¡£çŠ¶æ€ä¸ºå¤„ç†ä¸­
            document = await DocumentModel.get(id=document_id)
            document.status = "processing"
            await document.save()

            # 2. æå–æ–‡æ¡£å†…å®¹
            content = await document_parser.extract_content(file_path, file_type)
            if not content or not content.strip():
                raise Exception("æ–‡æ¡£å†…å®¹ä¸ºç©ºæˆ–æ— æ³•æå–")

            # æ›´æ–°æ–‡æ¡£å†…å®¹åˆ°æ•°æ®åº“
            document.content = content
            await document.save()

            # 3. æ™ºèƒ½åˆ†å—å¤„ç†
            chunks = self.text_splitter.split_text(content)
            if not chunks:
                raise Exception("æ–‡æ¡£åˆ†å—å¤±è´¥")

            # 4. åˆ›å»ºåˆ†å—è®°å½•
            chunk_objects = []
            for i, chunk_text in enumerate(chunks):
                chunk = await DocumentChunk.create(
                    document_id=document_id,
                    chunk_index=i,
                    content=chunk_text,
                    content_length=len(chunk_text),
                    metadata={"chunk_index": i}
                )
                chunk_objects.append(chunk)

            # 5. å­˜å‚¨åˆ°å‘é‡åº“
            if len(chunk_objects) > 0:
                # ä½¿ç”¨LangChainå‘é‡å­˜å‚¨æ·»åŠ æ–‡æ¡£ï¼ˆç›´æ¥ä½¿ç”¨å·²åˆ†å—çš„æ–‡æ¡£ï¼‰
                metadata = {
                    "filename": document.original_filename or document.filename,
                    "file_type": file_type,
                    "upload_time": document.upload_time.isoformat() if document.upload_time else None
                }

                await vector_search.add_documents_from_chunks(
                    document_id=document_id,
                    chunks=chunks,
                    chunk_objects=chunk_objects,
                    metadata=metadata
                )
            # 7. æ›´æ–°æ–‡æ¡£çŠ¶æ€ä¸ºå®Œæˆå¹¶è®¾ç½®å¤„ç†æ—¶é—´
            document.status = "completed"
            document.process_time = datetime.now()
            await document.save()

            return True

        except Exception as e:
            # æ›´æ–°æ–‡æ¡£çŠ¶æ€ä¸ºå¤±è´¥
            if document is None:
                document = await DocumentModel.get(id=document_id)
            document.status = "failed"
            document.error_message = str(e)
            await document.save()

            return False


# å…¨å±€å®ä¾‹ - å•ä¾‹æ¨¡å¼
try:
    document_parser = DocumentParser()
    document_processor = DocumentProcessor()
except Exception as e:
    raise Exception(f"æ–‡æ¡£å¤„ç†ç³»ç»Ÿåˆå§‹åŒ–å¤±è´¥: {e}")
